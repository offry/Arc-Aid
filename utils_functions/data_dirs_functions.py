from utils_functions.imports import *

import shutil
import os
import pickle

def find_diff_drawing_and_photo(drawing_dir, photo_dir):
    drawing_list, photo_list = [], []
    drawing_dir = os.path.join(os.getcwd(), drawing_dir)
    for label in os.listdir(drawing_dir):
        if label=="Sphinx" or label=="Duck":
            continue
        current_dir = os.path.join(drawing_dir, label)
        drawing_list.extend(os.listdir(current_dir))
    photo_dir = os.path.join(os.getcwd(), photo_dir)
    for label in os.listdir(photo_dir):
        if label=="Sphinx" or label=="Duck":
            continue
        current_dir = os.path.join(photo_dir, label)
        photo_list.extend(os.listdir(current_dir))
    drawing_list = [i.split('_')[0] for i in drawing_list]
    photo_list = [i.split('_')[0] for i in photo_list]
    not_in_photo = list(set(drawing_list)-set(photo_list))
    not_in_drawing = list(set(photo_list)-set(drawing_list))

    for i in not_in_photo:
        print("not in photo: {}".format(i))
    for i in not_in_drawing:
        print("not in drawing: {}".format(i))


def create_all_data_drawing(all_data_drawing, drawing_dir):
    drawing_dir = os.path.join(os.getcwd(), drawing_dir)
    all_data_drawing = os.path.join(os.getcwd(), all_data_drawing)
    if not os.path.isdir(all_data_drawing):
        os.mkdir(all_data_drawing)
    for label in os.listdir(drawing_dir):
        if label=="Sphinx" or label=="Duck":
            continue
        current_dir = os.path.join(drawing_dir, label)
        for file in os.listdir(current_dir):
            src = os.path.join(current_dir, file)
            dst = os.path.join(all_data_drawing, file)
            print(file)
            shutil.copy2(src, dst)


def create_all_data_photo(all_data_photo, photo_dir):
    photo_dir = os.path.join(os.getcwd(), photo_dir)
    all_data_photo = os.path.join(os.getcwd(), all_data_photo)
    if not os.path.isdir(all_data_photo):
        os.mkdir(all_data_photo)
    for label in os.listdir(photo_dir):
        if label=="Sphinx" or label=="Duck":
            continue
        current_dir = os.path.join(photo_dir, label)
        for file in os.listdir(current_dir):
            src = os.path.join(current_dir, file)
            dst = os.path.join(all_data_photo, file)
            print(file)
            shutil.copy2(src, dst)


def update_labels_excel(excel_name, images_dir):
    import openpyxl
    label_wrkbk = openpyxl.load_workbook(excel_name)
    label_sh = label_wrkbk.active
    labels_dir_list = os.listdir(images_dir)
    i = 2
    for label_dir in labels_dir_list:
        current_dir = os.path.join(images_dir, label_dir)
        for file in os.listdir(current_dir):
            image_name_obj = label_sh.cell(row=i, column=1)
            image_name_obj.value = file
            image_label_obj = label_sh.cell(row=i, column=2)
            image_label_obj.value = label_dir
            i+=1
    label_wrkbk.save(excel_name)


def update_photo_dirs(drawing_dir, photo_dir):
    drawing_dir_list = os.listdir(drawing_dir)
    photo_dir_list = os.listdir(photo_dir)
    new_all_data_dir = os.path.join(os.getcwd(),os.path.join("images_final", "drawing_base_all"))
    for dir in drawing_dir_list:
        dir_list = os.listdir(os.path.join(os.getcwd(), os.path.join(drawing_dir, dir)))
        for current_dir in dir_list:
            photo_current_dir = os.path.join(photo_dir, os.path.join(dir, current_dir))
            if not os.path.isdir(photo_current_dir):
                os.mkdir(photo_current_dir)
            for file in os.listdir(os.path.join(drawing_dir, os.path.join(dir, current_dir))):
                if "/" in file:
                    image_name_and_number = file.split("/")[-1].split("_")[0]
                else:
                    image_name_and_number = file.split("_")[0]
                for photo_file in os.listdir(new_all_data_dir):
                    if image_name_and_number in photo_file:
                        if "draw" not in photo_file:
                            break
                src = os.path.join(new_all_data_dir, photo_file)
                dst = os.path.join(photo_current_dir, photo_file)
                shutil.copy2(src, dst)
                print("copy {} to photo {}".format(image_name_and_number, dir + "_" + current_dir))

def update_photo_as_drawing_dir(drawing_dir, photo_dir):
    drawing_dir_list = os.listdir(drawing_dir)
    photo_dir_list = os.listdir(photo_dir)
    new_all_data_dir = os.path.join(os.getcwd(),os.path.join("images_final", "all_photo_as_drawing"))
    for dir in drawing_dir_list:
        dir_list = os.listdir(os.path.join(os.getcwd(), os.path.join(drawing_dir, dir)))
        for current_dir in dir_list:
            photo_current_dir = os.path.join(photo_dir, os.path.join(dir, current_dir))
            if not os.path.isdir(photo_current_dir):
                os.mkdir(photo_current_dir)
            for file in os.listdir(os.path.join(drawing_dir, os.path.join(dir, current_dir))):
                if "/" in file:
                    image_name_and_number = file.split("/")[-1].split("_")[0]
                else:
                    image_name_and_number = file.split("_")[0]
                for photo_file in os.listdir(new_all_data_dir):
                    if image_name_and_number in photo_file:
                        if "draw" not in photo_file:
                            break
                src = os.path.join(new_all_data_dir, photo_file)
                dst = os.path.join(photo_current_dir, photo_file)
                shutil.copy2(src, dst)
                print("copy {} to photo {}".format(image_name_and_number, dir + "_" + current_dir))

def create_labels_dict(labels_dict_name, args):
    if args.shapes_classification:
        task_type = "shapes"
        if not os.path.isdir("data/shapes"):
            os.mkdir("data/shapes")
        filename = "data/shapes/" + labels_dict_name

    if args.periods_classification:
        task_type = "periods"
        if not os.path.isdir("data/periods"):
            os.mkdir("data/periods")
        filename = "data/periods/" + labels_dict_name

    print("Creating labels_dict.bin")
    with open(filename, 'wb') as outfile:
        labels_dict = {}
        if "photo" in labels_dict_name:
            base_dir = "../../labels/" + task_type
        else:
            base_dir = "../../labels/" + task_type
        label_list = os.listdir(base_dir)
        for label in label_list:
            label_dir = os.path.join(base_dir, label)
            if "txt" in label:
                continue
            for image_name in os.listdir(label_dir):
                labels_dict[image_name] = str(label)
        pickle.dump(labels_dict, outfile)


def check_image_label(given_image_path, labels_dict_name):
    filename = labels_dict_name
    with open(filename, 'rb') as infile:
        labels_dict = pickle.load(infile)
    if "/" in given_image_path:
        image_name_and_number = given_image_path.split("/")[-1].split("_")[0]
    else:
        image_name_and_number = given_image_path.split("_")[0]
    # iterate through excel and display data
    for image in labels_dict.keys():
        image_split = image.split("_")
        image_after_split = ""
        for spl in image_split:
            image_after_split += spl + " "
        if image_name_and_number in image_after_split:
            return str(labels_dict[image])


def create_drawing_base_dirs(input_dir, photo_dir):
    photo_dir_list = os.listdir(photo_dir)
    input_dir_list = os.listdir(input_dir)

    labels_drawing_dir = os.path.join(os.path.join(os.getcwd(), "images_final"), "labeled_drawing")
    if not os.path.isdir(labels_drawing_dir):
        os.mkdir(labels_drawing_dir)

    for file in photo_dir_list:
        file_split_name = file.split("photoBase")
        drawing_file_name = [x for x in input_dir_list if file_split_name[0] in x and "draw" in x]
        if len(drawing_file_name)==1:
            drawing_label = check_image_label(file)
            if drawing_label is not None:
                label_dir = os.path.join(labels_drawing_dir, drawing_label)
                if not os.path.isdir(label_dir):
                    os.mkdir(label_dir)
                src_file = os.path.join(input_dir, drawing_file_name[0])
                dst_file = os.path.join(label_dir, drawing_file_name[0])
                if not os.path.isfile(dst_file):
                    if not os.path.isfile(os.path.join("images_final/drawing_not_use", drawing_label)):
                        shutil.copy2(src_file, dst_file)
                        print("copied {}".format(drawing_file_name[0]))


def calculate_how_many_images_in_each_label(labels_dict_name):
    filename = labels_dict_name
    with open(filename, 'rb') as infile:
        labels_dict = pickle.load(infile)
    labels_count_dict = {}
    for image in labels_dict.keys():
        if str(labels_dict[image]) in labels_count_dict:
            labels_count_dict[str(labels_dict[image])] = labels_count_dict[str(labels_dict[image])] + 1
        else:
            labels_count_dict[str(labels_dict[image])] = 1
    return labels_count_dict

def create_augmentations(dir):
    labels_dir_list = sorted(os.listdir(dir))
    for label in labels_dir_list:
        label_dir = os.path.join(dir, label)
        list_dir = sorted(os.listdir(label_dir))
        for i, file in zip(range(len(list_dir)), list_dir):
            file_name = os.path.join(label_dir, file)
            originalImage = cv2.imread(file_name)
            flipHorizontal = cv2.flip(originalImage, 1)
            cv2.imwrite(os.path.join(file_name.split("tif")[0] + "_flip.tif"), flipHorizontal)
            print(file_name + "_flip")
        list_dir = sorted(os.listdir(label_dir))
        for i, file in zip(range(len(list_dir)), list_dir):
            file_name = os.path.join(label_dir, file)
            originalImage = cv2.imread(file_name)
            if "draw" in dir:
                inverse = originalImage
            else:
                inverse =  cv2.bitwise_not(originalImage)
            cv2.imwrite(os.path.join(file_name.split("tif")[0] + "_inverse.tif"), inverse)
            print(file_name + "_inverse")

def create_photo_label_from_drawings(drawing_dir, photo_dir):
    new_all_data_dir = os.path.join(os.getcwd(), os.path.join("images_final", "all_image_base/all_image_base"))
    for file in os.listdir(os.path.join(drawing_dir)):
        if "/" in file:
            image_name_and_number = file.split("/")[-1].split("_")[0]
        else:
            image_name_and_number = file.split("_")[0]
        for photo_file in os.listdir(new_all_data_dir):
            if image_name_and_number in photo_file:
                break
        src = os.path.join(new_all_data_dir, photo_file)
        dst = os.path.join(photo_dir, photo_file)
        shutil.copy2(src, dst)
        print("copy {} to photo {}".format(image_name_and_number, photo_dir))
